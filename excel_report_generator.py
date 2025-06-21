"""
Модуль для генерации Excel отчетов
Excel Report Generator Module

Извлеченная часть из системы идентификации ГОСТ документов
Extracted part from GOST document identification system
"""

import logging
from datetime import datetime
from pathlib import Path
from typing import List, Dict, Any, Optional
import pandas as pd

logger = logging.getLogger(__name__)


class Document:
    """Модель нормативного документа / Document model"""

    def __init__(self,
                 doc_type: str,
                 number: str,
                 date: Optional[datetime] = None,
                 title: Optional[str] = None,
                 status: str = "статус неизвестен",
                 validation_source: Optional[str] = None,
                 validation_date: Optional[datetime] = None,
                 confidence: Optional[float] = None,
                 legal_api_status: Optional[str] = None,
                 legal_api_url: Optional[str] = None):
        self.type = doc_type
        self.number = number
        self.date = date
        self.title = title
        self.status = status
        self.validation_source = validation_source
        self.validation_date = validation_date
        self.confidence = confidence
        self.legal_api_status = legal_api_status
        self.legal_api_url = legal_api_url

    def to_excel_row(self) -> dict:
        """Преобразование в строку для Excel отчета / Convert to Excel row"""
        return {
            "Тип документа": self.type,
            "Номер": self.number,
            "Дата": self.date.strftime("%d.%m.%Y") if self.date else "",
            "Название": self.title or "",
            "Статус": self.status,
            "Источник проверки": self.validation_source or "",
            "Дата проверки": self.validation_date.strftime("%d.%m.%Y %H:%M") if self.validation_date else "",
            "Уверенность": f"{self.confidence:.2f}" if self.confidence else "",
            "Legal API статус": self.legal_api_status or "",
            "Legal API URL": self.legal_api_url or ""
        }


class ExcelReportGenerator:
    """Генератор Excel отчетов / Excel Report Generator"""

    def __init__(self):
        self.default_columns = [
            "№ п/п",
            "Тип документа",
            "Номер",
            "Дата",
            "Название",
            "Статус",
            "Источник проверки",
            "Дата проверки",
            "Уверенность",
            "Legal API статус",
            "Legal API URL"
        ]

    def create_report(self, report_data: Dict[str, Any], output_path: str) -> None:
        """
        Создание Excel отчета / Create Excel report

        Args:
            report_data: Данные для отчета / Report data
            output_path: Путь для сохранения файла / Output file path
        """
        try:
            documents = report_data.get('documents', [])

            if not documents:
                logger.warning("Нет документов для экспорта / No documents to export")
                return

            # Создаем DataFrame / Create DataFrame
            rows = []
            for i, doc in enumerate(documents, 1):
                row = doc.to_excel_row()
                row["№ п/п"] = i
                rows.append(row)

            df = pd.DataFrame(rows)

            # Переупорядочиваем колонки / Reorder columns
            available_columns = [col for col in self.default_columns if col in df.columns]
            extra_columns = [col for col in df.columns if col not in self.default_columns]
            df = df[available_columns + extra_columns]

            # Создаем Excel файл с несколькими листами / Create Excel file with multiple sheets
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                # Основной лист с документами / Main sheet with documents
                df.to_excel(writer, sheet_name='Документы', index=False)

                # Лист со статистикой / Statistics sheet
                stats_data = self._create_statistics(documents, report_data)
                stats_df = pd.DataFrame(list(stats_data.items()), columns=['Параметр', 'Значение'])
                stats_df.to_excel(writer, sheet_name='Статистика', index=False)

                # Лист с результатами поиска Legal API / Legal API search results sheet
                if any(doc.legal_api_status for doc in documents):
                    legal_api_data = self._create_legal_api_summary(documents)
                    legal_api_df = pd.DataFrame(legal_api_data)
                    legal_api_df.to_excel(writer, sheet_name='Legal API Results', index=False)

                # Форматирование основного листа / Format main sheet
                worksheet = writer.sheets['Документы']
                self._format_worksheet(worksheet, df)

            logger.info(f"Excel отчет создан / Excel report created: {output_path}")

        except Exception as e:
            logger.error(f"Ошибка при создании Excel отчета / Error creating Excel report: {str(e)}")
            raise

    def create_search_report(self, search_results: List[Dict], search_query: str, output_path: str) -> None:
        """
        Создание отчета по результатам поиска / Create search results report

        Args:
            search_results: Результаты поиска / Search results
            search_query: Поисковый запрос / Search query
            output_path: Путь для сохранения / Output path
        """
        try:
            # Преобразуем результаты поиска в документы
            documents = []
            for i, result in enumerate(search_results):
                doc = Document(
                    doc_type=result.get('type', 'Неизвестно'),
                    number=result.get('number', f'doc_{i}'),
                    date=result.get('date'),
                    title=result.get('title', ''),
                    status=result.get('status', 'Неизвестно'),
                    validation_source=result.get('source', ''),
                    validation_date=datetime.now(),
                    confidence=result.get('confidence', 0.0),
                    legal_api_status=result.get('legal_api_status'),
                    legal_api_url=result.get('legal_api_url')
                )
                documents.append(doc)

            # Создаем данные отчета
            report_data = {
                'documents': documents,
                'search_query': search_query,
                'search_date': datetime.now(),
                'total_results': len(search_results),
                'extraction_method': 'Legal Documents API Search',
                'processing_time': 0.0
            }

            self.create_report(report_data, output_path)

        except Exception as e:
            logger.error(f"Ошибка при создании отчета поиска: {e}")
            raise

    def _create_statistics(self, documents: List[Document], report_data: Dict[str, Any]) -> Dict[str, Any]:
        """Создание статистики для отчета / Create report statistics"""

        total_docs = len(documents)

        # Статистика по типам документов / Document type statistics
        type_stats = {}
        for doc in documents:
            type_stats[doc.type] = type_stats.get(doc.type, 0) + 1

        # Статистика по статусам / Status statistics
        status_stats = {}
        for doc in documents:
            status_stats[doc.status] = status_stats.get(doc.status, 0) + 1

        # Статистика по Legal API
        legal_api_stats = {}
        for doc in documents:
            if doc.legal_api_status:
                legal_api_stats[doc.legal_api_status] = legal_api_stats.get(doc.legal_api_status, 0) + 1

        # Средняя уверенность / Average confidence
        confidences = [doc.confidence for doc in documents if doc.confidence is not None]
        avg_confidence = sum(confidences) / len(confidences) if confidences else 0

        stats = {
            'Дата создания отчета': datetime.now().strftime('%d.%m.%Y %H:%M'),
            'Исходный файл': report_data.get('source_file', 'Не указан'),
            'Поисковый запрос': report_data.get('search_query', 'Не указан'),
            'Метод извлечения': report_data.get('extraction_method', 'Не указан'),
            'Время обработки (сек)': f"{report_data.get('processing_time', 0):.2f}",
            'Всего документов': total_docs,
            'Средняя уверенность': f"{avg_confidence:.2f}" if avg_confidence > 0 else "Не определена"
        }

        # Добавляем статистику по типам / Add type statistics
        for doc_type, count in type_stats.items():
            stats[f'Количество: {doc_type}'] = count

        # Добавляем статистику по статусам / Add status statistics
        for status, count in status_stats.items():
            stats[f'Статус: {status}'] = count

        # Добавляем статистику по Legal API
        for api_status, count in legal_api_stats.items():
            stats[f'Legal API: {api_status}'] = count

        # Добавляем метаданные текста если есть / Add text metadata if available
        text_metadata = report_data.get('text_metadata', {})
        if text_metadata:
            stats['Параграфов в документе'] = text_metadata.get('total_paragraphs', 0)
            stats['Таблиц в документе'] = text_metadata.get('total_tables', 0)
            stats['Размер файла (байт)'] = text_metadata.get('file_size', 0)

        return stats

    def _create_legal_api_summary(self, documents: List[Document]) -> List[Dict]:
        """Создание сводки по результатам Legal API"""
        summary = []

        for doc in documents:
            if doc.legal_api_status:
                summary.append({
                    'Документ': f"{doc.type} {doc.number}",
                    'Название': doc.title or '',
                    'Legal API статус': doc.legal_api_status,
                    'Базовый статус': doc.status,
                    'Совпадение статусов': 'Да' if doc.legal_api_status.lower() in doc.status.lower() else 'Нет',
                    'URL': doc.legal_api_url or ''
                })

        return summary

    def _format_worksheet(self, worksheet, df: pd.DataFrame) -> None:
        """Форматирование Excel листа / Format Excel worksheet"""
        try:
            from openpyxl.styles import Font, PatternFill, Alignment
            from openpyxl.utils import get_column_letter

            # Заголовки / Headers
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            center_alignment = Alignment(horizontal="center", vertical="center")

            # Применяем стили к заголовкам / Apply styles to headers
            for col_num, column in enumerate(df.columns, 1):
                cell = worksheet.cell(row=1, column=col_num)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_alignment

            # Автоширина колонок / Auto-width columns
            for column in worksheet.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)

                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass

                adjusted_width = min(max_length + 2, 50)  # Максимум 50 символов / Max 50 characters
                worksheet.column_dimensions[column_letter].width = adjusted_width

            # Заморозка первой строки / Freeze first row
            worksheet.freeze_panes = "A2"

        except ImportError:
            logger.warning(
                "openpyxl не поддерживает расширенное форматирование / openpyxl doesn't support advanced formatting")
        except Exception as e:
            logger.warning(f"Ошибка при форматировании листа / Error formatting worksheet: {str(e)}")


def create_sample_report():
    """Пример создания отчета / Example of creating a report"""

    # Создаем тестовые документы / Create sample documents
    documents = [
        Document(
            doc_type="ГОСТ",
            number="Р 7.0.5-2008",
            date=datetime(2008, 12, 15),
            title="Система стандартов по информации, библиотечному и издательскому делу. Библиографическая ссылка. Общие требования и правила составления",
            status="действует",
            validation_source="Росстандарт",
            validation_date=datetime.now(),
            confidence=0.95,
            legal_api_status="действующий",
            legal_api_url="https://gostexpert.ru/gost/7.0.5-2008"
        ),
        Document(
            doc_type="Приказ",
            number="1234",
            date=datetime(2023, 5, 20),
            title="О внесении изменений в технические регламенты",
            status="действует",
            validation_source="Минпромторг",
            validation_date=datetime.now(),
            confidence=0.87,
            legal_api_status="действующий",
            legal_api_url="https://pravo.gov.ru/document/1234"
        ),
        Document(
            doc_type="ГОСТ",
            number="15150-69",
            date=datetime(1969, 1, 1),
            title="Машины, приборы и другие технические изделия. Исполнения для различных климатических районов",
            status="действует",
            validation_source="Росстандарт",
            validation_date=datetime.now(),
            confidence=0.92,
            legal_api_status="действующий",
            legal_api_url="https://gostexpert.ru/gost/15150-69"
        )
    ]

    # Данные отчета / Report data
    report_data = {
        'documents': documents,
        'extraction_method': 'Legal Documents API + AI + Regex',
        'processing_time': 2.5,
        'source_file': 'sample_document.docx',
        'search_query': 'ГОСТ стандарты',
        'text_metadata': {
            'total_paragraphs': 150,
            'total_tables': 3,
            'file_size': 1024000
        }
    }

    # Создаем отчет / Create report
    generator = ExcelReportGenerator()
    output_path = "legal_documents_search_report.xlsx"
    generator.create_report(report_data, output_path)

    print(f"✅ Отчет создан / Report created: {output_path}")
    return output_path


def create_search_results_report(search_results: List[Dict], query: str, output_path: str = None) -> str:
    """
    Создание отчета по результатам поиска правовых документов
    Create report from legal documents search results

    Args:
        search_results: Результаты поиска
        query: Поисковый запрос
        output_path: Путь для сохранения (опционально)

    Returns:
        Путь к созданному файлу
    """
    if not output_path:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_path = f"search_results_{timestamp}.xlsx"

    generator = ExcelReportGenerator()
    generator.create_search_report(search_results, query, output_path)

    return output_path


if __name__ == "__main__":
    # Настройка логирования / Setup logging
    logging.basicConfig(level=logging.INFO)

    # Создаем пример отчета / Create sample report
    create_sample_report() 
